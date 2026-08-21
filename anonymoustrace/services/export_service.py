"""Export service for structured output."""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import Any

from anonymoustrace.models import ScanResult

logger = logging.getLogger(__name__)


class ExportService:
    """Handles JSON, CSV, XLSX, and TXT export of scan results."""

    def __init__(self, output_path: str | None = None, folder_output: str | None = None) -> None:
        self.output_path = Path(output_path) if output_path else None
        self.folder_output = Path(folder_output) if folder_output else None

    def _get_path(self, username: str, extension: str) -> Path:
        if self.output_path:
            path = self.output_path
        elif self.folder_output:
            path = self.folder_output / f"{username}.{extension}"
        else:
            path = Path(f"{username}.{extension}")
        return path

    def export_json(self, results: list[ScanResult], username: str) -> Path:
        """Export results as JSON. Returns the written path."""
        path = self._get_path(username, "json")

        data = {
            "username": username,
            "count": len(results),
            "results": [r.to_dict() for r in results],
        }

        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)

        logger.info("JSON exported to %s", path)
        return path

    def export_csv(self, results: list[ScanResult], username: str) -> Path:
        """Export results as CSV. Returns the written path."""
        path = self._get_path(username, "csv")

        fieldnames = ["username", "site", "detected", "confidence", "status_code", "error"]

        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for r in results:
                writer.writerow({
                    "username": r.username,
                    "site": r.site_name,
                    "detected": r.detected,
                    "confidence": r.confidence.value,
                    "status_code": r.status_code,
                    "error": r.error or "",
                })

        logger.info("CSV exported to %s", path)
        return path

    def export_xlsx(self, results: list[ScanResult], username: str) -> Path:
        """Export results as XLSX (Excel). Returns the written path."""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.warning("openpyxl not installed. Falling back to CSV.")
            return self.export_csv(results, username)

        path = self._get_path(username, "xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "OSINT Results"

        # Header row
        headers = ["Username", "Site", "Detected", "Confidence", "Status Code", "Error"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data rows
        for r in results:
            ws.append([
                r.username,
                r.site_name,
                "Yes" if r.detected else "No",
                r.confidence.value,
                r.status_code or "",
                r.error or "",
            ])

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

        logger.info("XLSX exported to %s", path)
        return path

    def export_txt(self, results: list[ScanResult], username: str) -> Path:
        """Export results as plain text. Returns the written path."""
        path = self._get_path(username, "txt")

        found = [r for r in results if r.detected]

        with open(path, "w", encoding="utf-8") as f:
            f.write(f"OSINT Reconnaissance Results for: {username}\n")
            f.write("=" * 50 + "\n\n")
            f.write(f"Total checked: {len(results)}\n")
            f.write(f"Found: {len(found)}\n")
            f.write(f"Not Found: {len(results) - len(found)}\n\n")
            
            f.write("FOUND PROFILES:\n")
            f.write("-" * 50 + "\n")
            for r in found:
                url = r.response_url or r.metadata.get("url", "N/A")
                f.write(f"[+] {r.site_name}: {url}\n")
            
            if not found:
                f.write("No profiles found.\n")

        logger.info("TXT exported to %s", path)
        return path

    def export(self, results: list[ScanResult], username: str, fmt: str) -> Path | None:
        """Export in the specified format."""
        if fmt == "json":
            return self.export_json(results, username)
        elif fmt == "csv":
            return self.export_csv(results, username)
        elif fmt == "xlsx":
            return self.export_xlsx(results, username)
        elif fmt == "txt":
            return self.export_txt(results, username)
        else:
            logger.warning("Unknown export format: %s", fmt)
            return None
